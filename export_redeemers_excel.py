import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

df = pd.read_csv('redeemer_customers_16859.csv')

with pd.ExcelWriter('redeemer_customers_16859.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Redeemer Customers')
    ws = writer.sheets['Redeemer Customers']

    # Column widths
    col_widths = [16, 16, 22, 20, 16, 18, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header style — dark navy
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Number formats per column (0-indexed from col 1)
    FMT_INR   = '\u20b9#,##0.00'
    FMT_NUM   = '#,##0'
    FMT_PCT   = '0.00'
    col_fmts  = [None, FMT_NUM, FMT_INR, FMT_NUM, FMT_PCT, FMT_INR, None, None]

    # Alternating row colors
    fill_even = PatternFill(start_color='EEF4FB', end_color='EEF4FB', fill_type='solid')

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for j, cell in enumerate(row):
            cell.alignment = Alignment(horizontal='left' if j == 0 else 'right', vertical='center')
            if col_fmts[j]:
                cell.number_format = col_fmts[j]
            if i % 2 == 0:
                cell.fill = fill_even

    # Summary row
    last = ws.max_row + 2
    summary_labels = ['TOTAL / AVG', df['Purchase_Count'].sum(), df['Total_Sale_Value'].sum(),
                      df['Total_Pts_Redeemed'].sum(), round(df['Redemption_Pct'].mean(), 2),
                      round(df['Avg_Bill_Value'].mean(), 2), '', '']
    summary_fmts = [None, FMT_NUM, FMT_INR, FMT_NUM, FMT_PCT, FMT_INR, None, None]
    gold_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    for j, (val, fmt) in enumerate(zip(summary_labels, summary_fmts), 1):
        cell = ws.cell(row=last, column=j, value=val)
        cell.font = Font(bold=True, size=11)
        cell.fill = gold_fill
        cell.alignment = Alignment(horizontal='left' if j == 1 else 'right', vertical='center')
        if fmt:
            cell.number_format = fmt

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

print(f'Done: redeemer_customers_16859.xlsx')
print(f'Rows: {len(df):,}')
print(f'Total Sale: Rs.{df["Total_Sale_Value"].sum():,.2f}')
print(f'Total Pts:  {df["Total_Pts_Redeemed"].sum():,.0f}')
